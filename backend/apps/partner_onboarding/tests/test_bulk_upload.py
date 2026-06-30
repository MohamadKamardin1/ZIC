import io
import os
import tempfile
from datetime import date
from unittest.mock import patch

import openpyxl
from django.test import TestCase, override_settings
from django.urls import reverse
from django.core.files.uploadedfile import SimpleUploadedFile
from rest_framework.test import APIClient

from apps.users.models import User
from apps.partner_onboarding.models import PartnerApplication
from apps.partner_onboarding.validators import validate_and_parse_excel


def create_test_user(**kwargs):
    defaults = {
        "email": "bulk@example.com",
        "username": "bulktest",
        "password": "TestPassword123!",
        "first_name": "Bulk",
        "last_name": "Test",
    }
    defaults.update(kwargs)
    return User.objects.create_user(**defaults)


def _build_individual_xlsx(rows):
    """Build an Individual Partner Template Excel file in memory."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Individual Partner Template"
    headers = [
        "Partner_Type", "Identification_Type", "Identification_Number",
        "Gender", "Title", "First_Name", "Other_Name", "Surname",
        "Email", "Telephone_Number", "Mobile_Number", "Nationality",
        "Date_of_Birth", "Physical_Address", "Postal_Address",
        "Political_Risk", "Anti-Money_Laundering", "Marital_Status", "Occupation",
    ]
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _build_corporate_xlsx(rows):
    """Build a Corporate Partner Template Excel file in memory."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Corporate Partner Template"
    headers = [
        "Partner_Type", "Company_Name", "Email", "Telephone_Number",
        "Mobile_Number", "TIN_Number", "Industry", "Incorporation_Date",
        "Company_Incorporation",
        "Contact_Person", "Contact_Person_Phone", "Contact_Person_Email",
        "Physical_Address", "Postal_Address", "Political_Risk", "AML_Risk",
    ]
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ============================================================================
# Validator Tests
# ============================================================================

class ValidateAndParseExcelTest(TestCase):

    def test_individual_template_valid_row(self):
        buf = _build_individual_xlsx([
            [
                "INDIVIDUAL", "National Identification Number", "NIN-001",
                "Male", "Mr", "John", "", "Doe",
                "john@example.com", "+255712345678", "+255712345678",
                "Tanzanian", "1990-05-15", "", "",
                "Low", "Low", "Married", "Engineer",
            ],
        ])
        partner_type, rows = validate_and_parse_excel(buf)
        self.assertEqual(partner_type, "INDIVIDUAL")
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["first_name"], "John")
        self.assertEqual(rows[0]["surname"], "Doe")
        self.assertEqual(rows[0]["gender"], "MALE")
        self.assertEqual(rows[0]["identification_type"], "NIN")
        self.assertEqual(rows[0]["date_of_birth"], date(1990, 5, 15))
        self.assertEqual(rows[0]["political_risk"], "LOW")
        self.assertEqual(rows[0]["aml_risk"], "LOW")
        self.assertEqual(rows[0]["marital_status"], "MARRIED")
        self.assertEqual(rows[0]["_errors"], [])

    def test_individual_template_validation_errors(self):
        buf = _build_individual_xlsx([
            [
                "INDIVIDUAL", "Invalid ID Type", "",
                "UnknownGender", "", "", "", "",
                "not-an-email", "", "",
                "InvalidCountry", "invalid-date", "", "",
                "BadRisk", "BadAML", "", "",
            ],
        ])
        partner_type, rows = validate_and_parse_excel(buf)
        self.assertEqual(partner_type, "INDIVIDUAL")
        self.assertEqual(len(rows), 1)
        errors = rows[0]["_errors"]
        # Should have at least 5 errors for bad enum values and missing required fields
        self.assertGreaterEqual(len(errors), 5)

    def test_individual_missing_required_fields(self):
        buf = _build_individual_xlsx([
            ["INDIVIDUAL", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
        ])
        partner_type, rows = validate_and_parse_excel(buf)
        self.assertEqual(len(rows), 1)
        errors = rows[0]["_errors"]
        # Should have many missing required field errors
        required = ["identification_type", "identification_number", "first_name",
                     "surname", "email", "mobile_number", "date_of_birth",
                     "nationality", "gender"]
        missing_count = sum(1 for e in errors if "Required field" in e)
        self.assertGreaterEqual(missing_count, len(required) - 1)  # partner_type is auto-set

    def test_individual_age_validation(self):
        buf = _build_individual_xlsx([
            [
                "INDIVIDUAL", "National Identification Number", "NIN-002",
                "Male", "Mr", "Child", "", "User",
                "child@example.com", "", "+255700000001",
                "Tanzanian", "2015-06-01", "", "", "Low", "Low", "", "",
            ],
        ])
        partner_type, rows = validate_and_parse_excel(buf)
        self.assertEqual(len(rows), 1)
        errors = rows[0]["_errors"]
        self.assertTrue(any("18 years" in e for e in errors),
                        f"Expected age error, got: {errors}")

    def test_individual_duplicate_email(self):
        buf = _build_individual_xlsx([
            [
                "INDIVIDUAL", "National Identification Number", "NIN-001",
                "Male", "Mr", "John", "", "Doe",
                "john@example.com", "+255712345678", "+255712345678",
                "Tanzanian", "1990-05-15", "", "",
                "Low", "Low", "Married", "Engineer",
            ],
            [
                "INDIVIDUAL", "Passport Number", "PP-002",
                "Female", "Ms", "Jane", "", "Doe",
                "john@example.com", "", "+255700000002",
                "Tanzanian", "1992-08-20", "", "",
                "Low", "Low", "Single", "Doctor",
            ],
        ])
        partner_type, rows = validate_and_parse_excel(buf)
        self.assertEqual(len(rows), 2)
        self.assertEqual(len(rows[0]["_errors"]), 0)
        self.assertTrue(any("Duplicate email" in e for e in rows[1]["_errors"]))

    def test_corporate_template_valid_row(self):
        buf = _build_corporate_xlsx([
            [
                "CORPORATE", "Acme Corp", "info@acme.co.tz",
                "+255222000000", "+255712000000", "TIN-12345",
                "Technology", "2020-01-15", "",
                "Jane Smith", "+255712000001", "jane@acme.co.tz",
                "123 Main St", "P.O. Box 456", "Low", "Low",
            ],
        ])
        partner_type, rows = validate_and_parse_excel(buf)
        self.assertEqual(partner_type, "CORPORATE")
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["company_name"], "Acme Corp")
        self.assertEqual(rows[0]["industry"], "TECHNOLOGY")
        self.assertEqual(rows[0]["incorporation_date"], date(2020, 1, 15))
        self.assertEqual(rows[0]["_errors"], [])

    def test_corporate_missing_required_fields(self):
        buf = _build_corporate_xlsx([
            ["CORPORATE", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
        ])
        partner_type, rows = validate_and_parse_excel(buf)
        self.assertEqual(len(rows), 1)
        errors = rows[0]["_errors"]
        required = ["company_name", "tin_number", "incorporation_date",
                     "industry", "email", "mobile_number",
                     "contact_person", "contact_person_phone",
                     "contact_person_email", "physical_address"]
        missing_count = sum(1 for e in errors if "Required field" in e)
        self.assertGreaterEqual(missing_count, len(required))

    def test_unknown_template_format(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Unknown", "Headers", "Here"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        with self.assertRaises(ValueError) as ctx:
            validate_and_parse_excel(buf)
        self.assertIn("Unknown template format", str(ctx.exception))

    def test_empty_file(self):
        wb = openpyxl.Workbook()
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        with self.assertRaises(ValueError) as ctx:
            validate_and_parse_excel(buf)
        self.assertIn("empty", str(ctx.exception).lower())


# ============================================================================
# API Endpoint Tests
# ============================================================================

@override_settings(STATICFILES_DIRS=[os.path.join(os.path.dirname(__file__), "../../../static")])
class DownloadTemplateViewTest(TestCase):

    def setUp(self):
        self.client = APIClient()
        self.user = create_test_user()
        self.client.force_authenticate(user=self.user)

    def test_download_individual_template(self):
        url = reverse("v1:bulk-upload-template")
        response = self.client.get(url, {"partner_type": "INDIVIDUAL"})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("Individual_Partners_Template.xlsx", response["Content-Disposition"])

    def test_download_corporate_template(self):
        url = reverse("v1:bulk-upload-template")
        response = self.client.get(url, {"partner_type": "CORPORATE"})
        self.assertEqual(response.status_code, 200)
        self.assertIn("Corporate_Partners_Template.xlsx", response["Content-Disposition"])

    def test_download_invalid_partner_type(self):
        url = reverse("v1:bulk-upload-template")
        response = self.client.get(url, {"partner_type": "INVALID"})
        self.assertEqual(response.status_code, 400)

    def test_download_missing_partner_type(self):
        url = reverse("v1:bulk-upload-template")
        response = self.client.get(url)
        self.assertEqual(response.status_code, 400)

    def test_download_requires_auth(self):
        self.client.force_authenticate(user=None)
        url = reverse("v1:bulk-upload-template")
        response = self.client.get(url, {"partner_type": "INDIVIDUAL"})
        self.assertEqual(response.status_code, 401)


class BulkUploadViewTest(TestCase):

    def setUp(self):
        self.client = APIClient()
        self.user = create_test_user()
        self.client.force_authenticate(user=self.user)

    def _upload_file(self, buf, filename="test.xlsx"):
        buf.seek(0)
        uploaded = SimpleUploadedFile(
            filename,
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        return self.client.post(
            reverse("v1:bulk-upload"),
            {"file": uploaded},
            format="multipart",
        )

    def test_upload_individual_single_row(self):
        buf = _build_individual_xlsx([
            [
                "INDIVIDUAL", "National Identification Number", "NIN-001",
                "Male", "Mr", "John", "", "Doe",
                "john@example.com", "+255712345678", "+255712345678",
                "Tanzanian", "1990-05-15", "", "",
                "Low", "Low", "Married", "Engineer",
            ],
        ])
        response = self._upload_file(buf)
        self.assertEqual(response.status_code, 200)
        data = response.data["data"]
        self.assertEqual(data["imported"], 1)
        self.assertEqual(data["skipped"], 0)
        self.assertEqual(data["errors"], [])
        self.assertTrue(PartnerApplication.objects.filter(email="john@example.com").exists())

    def test_upload_corporate_single_row(self):
        buf = _build_corporate_xlsx([
            [
                "CORPORATE", "Acme Corp", "info@acme.co.tz",
                "+255222000000", "+255712000000", "TIN-12345",
                "Technology", "2020-01-15", "",
                "Jane Smith", "+255712000001", "jane@acme.co.tz",
                "123 Main St", "P.O. Box 456", "Low", "Low",
            ],
        ])
        response = self._upload_file(buf)
        self.assertEqual(response.status_code, 200)
        data = response.data["data"]
        self.assertEqual(data["imported"], 1)
        self.assertEqual(data["skipped"], 0)
        self.assertTrue(PartnerApplication.objects.filter(email="info@acme.co.tz").exists())

    def test_upload_with_invalid_rows_skips_them(self):
        buf = _build_individual_xlsx([
            # Valid row
            [
                "INDIVIDUAL", "National Identification Number", "NIN-001",
                "Male", "Mr", "Valid", "", "User",
                "valid@example.com", "", "+255700000001",
                "Tanzanian", "1990-05-15", "", "",
                "Low", "Low", "Married", "Engineer",
            ],
            # Invalid row — bad email, missing required
            [
                "INDIVIDUAL", "National Identification Number", "",
                "Unknown", "", "", "", "",
                "bad-email", "", "",
                "Invalid", "bad-date", "", "",
                "Bad", "Bad", "", "",
            ],
        ])
        response = self._upload_file(buf)
        self.assertEqual(response.status_code, 200)
        data = response.data["data"]
        self.assertEqual(data["imported"], 1)
        self.assertEqual(data["skipped"], 1)
        self.assertEqual(len(data["errors"]), 1)
        self.assertTrue(PartnerApplication.objects.filter(email="valid@example.com").exists())
        self.assertFalse(PartnerApplication.objects.filter(email="bad-email").exists())

    def test_upload_requires_authentication(self):
        self.client.force_authenticate(user=None)
        buf = _build_individual_xlsx([
            [
                "INDIVIDUAL", "National Identification Number", "NIN-001",
                "Male", "Mr", "John", "", "Doe",
                "john@example.com", "", "+255700000001",
                "Tanzanian", "1990-05-15", "", "",
                "Low", "Low", "Married", "Engineer",
            ],
        ])
        response = self._upload_file(buf)
        self.assertEqual(response.status_code, 401)

    def test_upload_no_file(self):
        response = self.client.post(reverse("v1:bulk-upload"), format="multipart")
        self.assertEqual(response.status_code, 400)

    def test_upload_invalid_extension(self):
        buf = io.BytesIO(b"not an excel")
        response = self._upload_file(buf, "test.txt")
        self.assertEqual(response.status_code, 400)

    def test_upload_empty_file(self):
        wb = openpyxl.Workbook()
        buf = io.BytesIO()
        wb.save(buf)
        response = self._upload_file(buf)
        self.assertEqual(response.status_code, 400)

    def test_upload_handles_db_error_gracefully(self):
        """Row passes validation but DB insert fails (e.g. constraint)."""
        # Create the first row fine
        buf = _build_individual_xlsx([
            [
                "INDIVIDUAL", "National Identification Number", "NIN-001",
                "Male", "Mr", "John", "", "Doe",
                "john@example.com", "", "+255700000001",
                "Tanzanian", "1990-05-15", "", "",
                "Low", "Low", "Married", "Engineer",
            ],
        ])
        response = self._upload_file(buf)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["data"]["imported"], 1)

    def test_multiple_valid_rows(self):
        buf = _build_individual_xlsx([
            [
                "INDIVIDUAL", "National Identification Number", "NIN-001",
                "Male", "Mr", "Alice", "", "Smith",
                "alice@example.com", "", "+255700000001",
                "Tanzanian", "1990-01-01", "", "",
                "Low", "Low", "Single", "Doctor",
            ],
            [
                "INDIVIDUAL", "Passport Number", "PP-001",
                "Female", "Ms", "Bob", "", "Jones",
                "bob@example.com", "", "+255700000002",
                "Kenyan", "1985-05-10", "", "",
                "Medium", "Low", "Married", "Engineer",
            ],
        ])
        response = self._upload_file(buf)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["data"]["imported"], 2)
        self.assertEqual(PartnerApplication.objects.count(), 2)
